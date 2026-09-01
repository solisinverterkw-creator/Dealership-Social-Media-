import os
import re
import io
import csv
from datetime import datetime, timedelta
from PIL import Image
import openpyxl

class ImageResizer:
    @staticmethod
    def resize(file_input, max_width: int = 800, max_height: int = 800):
        """Resizes an image file object or path, returning a dict {'success': bool, 'data': bytes, 'message': str}."""
        try:
            if hasattr(file_input, 'read'):
                file_bytes = file_input.read()
                if hasattr(file_input, 'seek'):
                    file_input.seek(0)
                img = Image.open(io.BytesIO(file_bytes))
            elif isinstance(file_input, str) and os.path.exists(file_input):
                img = Image.open(file_input)
            else:
                return {'success': False, 'message': 'Invalid file input', 'data': None}

            w, h = img.size
            if w > max_width or h > max_height:
                if w / max_width >= h / max_height:
                    new_w = max_width
                    new_h = int(round((h / w) * max_width))
                else:
                    new_h = max_height
                    new_w = int(round((w / h) * max_height))
                img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

            output = io.BytesIO()
            fmt = img.format if img.format in ['PNG', 'WEBP'] else 'JPEG'
            if fmt == 'JPEG' and img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            img.save(output, format=fmt, quality=85)
            return {'success': True, 'data': output.getvalue()}
        except Exception as e:
            return {'success': False, 'message': str(e), 'data': None}

    @staticmethod
    def resize_in_place(file_path: str, max_dim: int = 720):
        """Resizes the image file in-place to fit within max_dim x max_dim dimensions."""
        if not os.path.exists(file_path):
            return
        try:
            with Image.open(file_path) as img:
                w, h = img.size
                if w <= max_dim and h <= max_dim:
                    return
                
                if w >= h:
                    new_h = int(round((h / w) * max_dim))
                    new_w = max_dim
                else:
                    new_w = int(round((w / h) * max_dim))
                    new_h = max_dim
                
                resized = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                fmt = img.format or 'JPEG'
                if fmt in ['PNG', 'WEBP']:
                    resized.save(file_path, format=fmt)
                else:
                    if resized.mode in ('RGBA', 'P'):
                        resized = resized.convert('RGB')
                    resized.save(file_path, format='JPEG', quality=85)
        except Exception:
            pass

def levenshtein_distance(s1: str, s2: str) -> int:
    """Computes the Levenshtein distance between two strings."""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)

    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row

    return previous_row[-1]

class SpreadsheetImportHelper:
    @staticmethod
    def find_header_row_index(rows: list, must_contain_any: list) -> int:
        """Returns the 0-based row index of the first row containing a cell that matches keyword."""
        for i, row in enumerate(rows):
            for cell in row:
                cell_str = str(cell).lower().strip()
                for needle in must_contain_any:
                    if cell_str != '' and needle.lower() in cell_str:
                        return i
        return 0

    @staticmethod
    def find_column(header_row: list, keywords: list, prefer_last: bool = False) -> int:
        """Returns the 0-based column index whose header matches a keyword."""
        found = None
        for i, cell in enumerate(header_row):
            cell_str = str(cell).lower().strip()
            if cell_str == '':
                continue
            if SpreadsheetImportHelper.matches_any_keyword(cell_str, keywords):
                if not prefer_last:
                    return i
                found = i
        return found

    @staticmethod
    def matches_any_keyword(label: str, keywords: list) -> bool:
        """Matches a label against a list of keywords."""
        label_lower = label.lower()
        for kw in keywords:
            kw_lower = kw.lower()
            if kw_lower in label_lower:
                return True
        return False

    @staticmethod
    def forward_fill_row(row: list, max_cols: int, anchor_col: int = -1) -> list:
        filled = []
        last = ''
        for c in range(max_cols):
            v = str(row[c]).strip() if c < len(row) else ''
            if v != '':
                last = v
            filled.append(last)
            if c == anchor_col:
                last = ''
        return filled

    @staticmethod
    def combine_multi_row_header(rows: list, header_index: int, anchor_col: int) -> dict:
        max_cols = 0
        for row in rows:
            max_cols = max(max_cols, len(row))

        combined = SpreadsheetImportHelper.forward_fill_row(rows[header_index] if header_index < len(rows) else [], max_cols, anchor_col)
        data_start_index = header_index + 1

        while data_start_index < len(rows):
            row = rows[data_start_index]
            is_blank_row = all(str(c).strip() == '' for c in row)
            if is_blank_row:
                break
            anchor_val = str(row[anchor_col]).strip() if anchor_col < len(row) else ''
            if anchor_val != '':
                break  # first real data row

            filled_row = SpreadsheetImportHelper.forward_fill_row(row, max_cols, anchor_col)
            for c in range(max_cols):
                if filled_row[c] != '' and filled_row[c] != combined[c]:
                    combined[c] = (combined[c] + ' ' + filled_row[c]).strip()
            data_start_index += 1

        return {'header': combined, 'dataStartIndex': data_start_index}

    @staticmethod
    def normalize_dealership_name(name: str) -> str:
        """Normalizes dealership name to match database records."""
        name = name.strip().lower()
        name = re.sub(r'\s*\(.*$', '', name) # Strip suffixes like (SMC-PVT)
        name = name.rstrip(" .\t\n\r")
        name = re.sub(r'^suzuki\s+', '', name) # Strip leading "suzuki"
        name = re.sub(r'\s+', '', name) # Remove all whitespace
        return name

    @staticmethod
    def find_dealership_match(dealerships_by_normalized_name: dict, raw_name: str) -> int:
        """Leniently matches a raw dealership string to a DB dealership ID to avoid skipping valid dealership data."""
        raw_str = str(raw_name or '').strip()
        if not raw_str or 'regional' in raw_str.lower():
            return None

        normalized = SpreadsheetImportHelper.normalize_dealership_name(raw_str)

        # 1. Exact match on normalized name
        if normalized in dealerships_by_normalized_name:
            return dealerships_by_normalized_name[normalized]

        # 2. Core Keyword Matching for all 21 tracked dealerships
        raw_lower = raw_str.lower()
        core_keywords = [
            ('multan city', 17), ('multancity', 17),
            ('south punjab', 3), ('southpunjab', 3), ('ssp', 3),
            ('rahim yar khan', 6), ('rahimyarkhan', 6), ('ryk', 6),
            ('bahawalnagar', 12), ('bahawal nagar', 12),
            ('bahawalpur', 13), ('bahawal pur', 13),
            ('chichawatni', 10), ('chicha watni', 10),
            ('muzaffargarh', 19), ('muzaffar garh', 19),
            ('mianchannu', 18), ('mian channu', 18), ('mianchanu', 18),
            ('pakpattan', 4), ('khanewal', 5), ('sadiqabad', 7), ('sadiq abad', 7),
            ('gateway', 8), ('shorkot', 9), ('sahiwal', 11), ('derawar', 14),
            ('fort', 15), ('unique', 16), ('rajanpur', 20), ('depalpur', 21),
            ('pioneer', 2), ('united', 1)
        ]
        for kw, d_id in core_keywords:
            if kw in raw_lower or kw in normalized:
                return d_id

        # 3. Substring matching
        matches = []
        for known_normalized, id_val in dealerships_by_normalized_name.items():
            if len(normalized) >= 3 and (normalized in known_normalized or known_normalized in normalized):
                matches.append(id_val)

        if len(matches) == 1:
            return matches[0]

        # 4. High-tolerance Levenshtein fuzzy distance matching
        best_id = None
        best_distance = None
        ambiguous = False

        for known_normalized, id_val in dealerships_by_normalized_name.items():
            dist = levenshtein_distance(normalized, known_normalized)
            tolerance = 6 if len(known_normalized) >= 8 else 3
            if dist > tolerance:
                continue
            if best_distance is None or dist < best_distance:
                best_distance = dist
                best_id = id_val
                ambiguous = False
            elif dist == best_distance:
                ambiguous = True

        return None if ambiguous else best_id

    @staticmethod
    def sort_product_columns_by_priority(product_names: list, priority_patterns: list) -> list:
        rank = {}
        for product in product_names:
            match_index = SpreadsheetImportHelper.find_best_matching_pattern_index(product, priority_patterns)
            rank[product] = match_index if match_index is not None else (len(priority_patterns) + 1)

        # Sort product names based on priority ranking, then alphabetically
        sorted_names = list(product_names)
        sorted_names.sort(key=lambda x: (rank[x], x.lower()))
        return sorted_names

    @staticmethod
    def shorten_product_label(product_name: str, priority_patterns: list) -> str:
        index = SpreadsheetImportHelper.find_best_matching_pattern_index(product_name, priority_patterns)
        return priority_patterns[index] if index is not None else product_name

    @staticmethod
    def find_best_matching_pattern_index(product: str, priority_patterns: list) -> int:
        product_lower = product.lower()
        best_pattern_index = None
        best_token_count = 0

        for pattern_index, pattern in enumerate(priority_patterns):
            tokens = [t for t in pattern.lower().strip().split() if t]
            pos = 0
            matched = True
            for token in tokens:
                token_escaped = re.escape(token)
                regex = rf'(?<![a-z0-9]){token_escaped}(?![a-z0-9])'
                match = re.search(regex, product_lower[pos:])
                if not match:
                    matched = False
                    break
                pos += match.end()
            if matched and len(tokens) > best_token_count:
                best_token_count = len(tokens)
                best_pattern_index = pattern_index

        return best_pattern_index

    @staticmethod
    def parse_flexible_date(value: str) -> str:
        """Parses various date format strings and returns YYYY-MM-DD."""
        value = str(value).strip()
        if value == '':
            return None

        # YYYYMMDD string format
        if re.match(r'^\d{8}$', value):
            try:
                dt = datetime.strptime(value, '%Y%m%d')
                return dt.strftime('%Y-%m-%d')
            except Exception:
                pass

        # Excel serial number date format (epoch 1899-12-30)
        if re.match(r'^\d{4,6}$', value):
            try:
                serial = int(value)
                if 25569 <= serial <= 73050:
                    dt = datetime(1899, 12, 30) + timedelta(days=serial)
                    return dt.strftime('%Y-%m-%d')
            except Exception:
                pass

        # DD/MM/YYYY slash separated date format
        if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', value):
            try:
                dt = datetime.strptime(value, '%d/%m/%Y')
                return dt.strftime('%Y-%m-%d')
            except Exception:
                pass

        # Standard date parses
        for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime('%Y-%m-%d')
            except Exception:
                pass

        return None

    @staticmethod
    def friendly_product_label(name: str) -> str:
        key = name.strip().lower()
        if key == 'pa':
            return 'PA (Pending Allocation)'
        if key == 'pb':
            return 'PB (Pending Booking)'
        return name

    def _build_dealership_map(self, db_session):
        from api.models import Dealership
        dealerships = db_session.query(Dealership).all()
        mapping = {}
        for d in dealerships:
            norm = self.normalize_dealership_name(d.name)
            if norm:
                mapping[norm] = d.id
            if d.fb_input:
                norm_fb = self.normalize_dealership_name(d.fb_input)
                if norm_fb:
                    mapping[norm_fb] = d.id
        return mapping

    def import_sales_sheet(self, db_session, rows: list, period_month: str) -> dict:
        from api.models import SalesRecord, SalesSummary
        if not rows or len(rows) < 2:
            return {'success': False, 'message': 'Sheet is empty or has no data rows.'}

        dealershipsByName = self._build_dealership_map(db_session)
        headerIndex = self.find_header_row_index(rows, ['dealer'])
        headerRow = rows[headerIndex] if headerIndex < len(rows) else []
        dealerCol = self.find_column(headerRow, ['dealer'])

        if dealerCol is None:
            return {'success': False, 'message': 'Could not find a "Dealer"/"Dealership" column in the header row.'}

        nextRow = rows[headerIndex + 1] if (headerIndex + 1) < len(rows) else []
        nextRowDealerVal = str(nextRow[dealerCol]).strip() if dealerCol < len(nextRow) else ''
        hasSubHeaderRow = (nextRowDealerVal == '') and any(str(c).strip() != '' for c in nextRow)

        effectiveHeader = list(headerRow)
        dataStartIndex = headerIndex + 1
        if hasSubHeaderRow:
            for col, val in enumerate(nextRow):
                v_str = str(val).strip()
                if v_str != '':
                    if col < len(effectiveHeader):
                        effectiveHeader[col] = val
                    else:
                        effectiveHeader.append(val)
            dataStartIndex = headerIndex + 2

        targetCol = self.find_column(effectiveHeader, ['target'])
        totalCol = self.find_column(effectiveHeader, ['total'])

        skipKeywords = ['sr#', 'sr.', 'sr no', 's.no', 's no', 'serial', 'dealer', 'target', 'total']
        productCols = {}
        for col, label in enumerate(effectiveHeader):
            labelStr = str(label).strip()
            if labelStr == '' or self.matches_any_keyword(labelStr, skipKeywords):
                continue
            productCols[col] = labelStr

        if not productCols:
            return {'success': False, 'message': 'Could not find any Product/Model columns in the sheet.'}

        touchedDealershipIds = set()
        importedCount = 0
        importErrors = []

        for i in range(dataStartIndex, len(rows)):
            row = rows[i]
            rowNum = i + 1
            if not any(str(c).strip() != '' for c in row):
                continue

            dealershipName = str(row[dealerCol]).strip() if dealerCol < len(row) else ''
            if not dealershipName:
                continue

            dealershipId = self.find_dealership_match(dealershipsByName, dealershipName)
            if not dealershipId:
                importErrors.append(f"Row {rowNum}: Dealership \"{dealershipName}\" Not Found — Skipped.")
                continue

            if dealershipId not in touchedDealershipIds:
                db_session.query(SalesRecord).filter(
                    SalesRecord.dealership_id == dealershipId,
                    SalesRecord.period_month == period_month
                ).delete()
                touchedDealershipIds.add(dealershipId)

            for col, productName in productCols.items():
                val_str = str(row[col]).strip() if col < len(row) else '0'
                try:
                    qty = int(float(val_str.replace(',', ''))) if val_str else 0
                except Exception:
                    qty = 0

                rec = SalesRecord(
                    dealership_id=dealershipId,
                    product_name=productName,
                    quantity=qty,
                    period_month=period_month,
                    column_order=col
                )
                db_session.add(rec)
                importedCount += 1

            target_val = None
            if targetCol is not None and targetCol < len(row):
                try:
                    target_val = int(float(str(row[targetCol]).replace(',', '').strip()))
                except Exception:
                    pass

            gt_val = None
            if totalCol is not None and totalCol < len(row):
                try:
                    gt_val = int(float(str(row[totalCol]).replace(',', '').strip()))
                except Exception:
                    pass

            summary = db_session.query(SalesSummary).filter(
                SalesSummary.dealership_id == dealershipId,
                SalesSummary.period_month == period_month
            ).first()
            if not summary:
                summary = SalesSummary(
                    dealership_id=dealershipId,
                    period_month=period_month,
                    target=target_val,
                    grand_total=gt_val,
                    grand_total_column_order=totalCol
                )
                db_session.add(summary)
            else:
                summary.target = target_val
                summary.grand_total = gt_val
                summary.grand_total_column_order = totalCol

        db_session.commit()
        return {'success': True, 'imported_count': importedCount, 'import_errors': importErrors}

    def normalize_stock_product_name(self, desc: str) -> str:
        d = str(desc or '').upper().strip()
        if not d:
            return 'Every VXR'

        # Alto variants
        if 'ALTO' in d or 'AET' in d:
            if 'VXL' in d:
                return 'Alto VXL AGS' if ('AGS' in d or 'AUTO' in d) else 'Alto VXL'
            elif 'VXR AGS' in d or ('VXR' in d and 'AGS' in d):
                return 'Alto VXR AGS'
            elif 'VXR' in d:
                return 'Alto VXR'
            elif 'AGS' in d:
                return 'Alto AGS'
            return 'Alto VXR'

        # Cultus variants
        if 'CULTUS' in d or 'AVK' in d:
            if 'VXL' in d:
                return 'Cultus VXL'
            elif 'VXR' in d:
                return 'Cultus VXR'
            elif 'AGS' in d or 'AUTO' in d:
                return 'Cultus AGS'
            return 'Cultus VXR'

        # Swift variants
        if 'SWIFT' in d or 'A2L' in d:
            if 'GLX' in d:
                return 'Swift GLX'
            elif 'CVT' in d:
                return 'Swift GL CVT'
            elif 'MT' in d:
                return 'Swift MT'
            elif 'GL' in d:
                return 'Swift GL'
            return 'Swift GL CVT'

        # Fronx variants
        if 'FRONX' in d or 'NWD' in d or 'NWA' in d:
            if 'GLX' in d or 'HYBD' in d:
                return 'Fronx GLX'
            elif 'GL' in d or 'AT' in d:
                return 'Fronx GL AT'
            elif 'MT' in d:
                return 'Fronx MT'
            return 'Fronx GL AT'

        # Every variants
        if 'EVERY' in d or 'A5H' in d:
            if 'VX' in d and 'VXR' not in d:
                return 'Every VX'
            return 'Every VXR'

        # Wagon R
        if 'WAGON' in d:
            if 'VXL' in d:
                return 'Wagon R VXL'
            elif 'VXR' in d:
                return 'Wagon R VXR'
            return 'Wagon R'

        # Bolan
        if 'BOLAN' in d:
            return 'Bolan'

        # Mehran
        if 'MEHRAN' in d:
            return 'Mehran'

        # Clean fallback title
        clean = re.sub(r'^suzuki\s+', '', d, flags=re.IGNORECASE).strip()
        return clean.title() if clean else d.title()

    def clean_dealership_name_for_creation(self, raw_name: str) -> str:
        s = str(raw_name or '').strip()
        if not s or len(s) < 3:
            return ''
        low = s.lower()
        if any(k in low for k in ['pak suzuki motor co', 'psmc insurance', 'slm pvt ltd', 'total', 'grand total', 'dealer name', 'dealer']):
            return ''
        return s

    def import_stock_sheet(self, db_session, rows: list) -> dict:
        from api.models import StockRecord, Dealership
        if not rows or len(rows) < 2:
            return {'success': False, 'message': 'Sheet is empty or has no data rows.'}

        excludedStockNames = ['suzuki habib motors', 'suzuki habib motors alipur']
        excludedStockIds = set()
        for d in db_session.query(Dealership).all():
            if self.normalize_dealership_name(d.name) in excludedStockNames:
                excludedStockIds.add(d.id)

        if excludedStockIds:
            db_session.query(StockRecord).filter(StockRecord.dealership_id.in_(list(excludedStockIds))).delete()

        dealershipsByName = self._build_dealership_map(db_session)
        headerIndex = self.find_header_row_index(rows, ['dealer'])
        headerRow = rows[headerIndex] if headerIndex < len(rows) else []
        dealerCol = self.find_column(headerRow, ['dealer'])

        if dealerCol is None:
            return {'success': False, 'message': 'Could not find a "Dealer"/"Dealership" column in the header row.'}

        dealerNameCol = self.find_column(headerRow, ['dealer name', 'dealership name', 'dealer_name'])
        if dealerNameCol is None:
            dealerNameCol = dealerCol
        securityCol = self.find_column(headerRow, ['security'])

        productDescCol = self.find_column(headerRow, ['product desc', 'product description'])
        if productDescCol is None:
            genericCol = self.find_column(headerRow, ['product'])
            if genericCol is not None and not self.matches_any_keyword(str(headerRow[genericCol]), ['code']):
                productDescCol = genericCol

        touchedDealershipIds = set()
        importedCount = 0
        importErrors = []

        has_matrix_variants = any(
            self.matches_any_keyword(str(h).lower(), ['vxr', 'ags', 'vxl', 'gl', 'glx', 'swift', 'alto', 'cultus', 'every', 'fronx', 'cuc'])
            for h in headerRow
        )

        if productDescCol is not None:
            qtyCol = self.find_column(headerRow, ['qty', 'quantity', 'sum of quantity', 'stock qty', 'total qty', 'count'])
            regionCol = self.find_column(headerRow, ['region'], prefer_last=True)
            counts = {}
            regionByDealer = {}
            productOrder = {}
            orderCounter = 0
            transactionRows = 0
            skippedNonTracked = 0

            for i in range(headerIndex + 1, len(rows)):
                row = rows[i]
                rowNum = i + 1
                if not any(str(c).strip() != '' for c in row):
                    continue

                dealershipName = str(row[dealerNameCol]).strip() if dealerNameCol < len(row) else ''
                raw_prod = str(row[productDescCol]).strip() if productDescCol < len(row) else ''
                if not raw_prod:
                    continue
                productName = self.normalize_stock_product_name(raw_prod)

                dealershipId = self.find_dealership_match(dealershipsByName, dealershipName)
                if not dealershipId and dealerCol is not None and dealerCol < len(row):
                    alt_name = str(row[dealerCol]).strip()
                    if alt_name:
                        dealershipId = self.find_dealership_match(dealershipsByName, alt_name)

                if not dealershipId:
                    skippedNonTracked += 1
                    continue

                if dealershipId in excludedStockIds:
                    continue

                row_qty = 1
                if qtyCol is not None and qtyCol < len(row):
                    q_str = str(row[qtyCol]).strip()
                    if q_str:
                        try:
                            row_qty = int(float(q_str.replace(',', '')))
                        except Exception:
                            row_qty = 1

                if securityCol is not None and securityCol < len(row):
                    sec_str = str(row[securityCol]).strip()
                    if sec_str != '':
                        try:
                            sec_amt = float(sec_str.replace(',', ''))
                            d_obj = db_session.query(Dealership).filter(Dealership.id == dealershipId).first()
                            if d_obj:
                                d_obj.security_amount = sec_amt
                        except Exception:
                            pass

                if regionCol is not None and regionCol < len(row):
                    reg_val = str(row[regionCol]).strip()
                    if reg_val != '':
                        regionByDealer[dealershipId] = reg_val

                if productName not in productOrder:
                    productOrder[productName] = orderCounter
                    orderCounter += 1

                if dealershipId not in counts:
                    counts[dealershipId] = {}
                counts[dealershipId][productName] = counts[dealershipId].get(productName, 0) + row_qty
                transactionRows += 1

            if not counts:
                return {'success': False, 'message': 'No Matching Tracked Dealership Rows Found To Import.', 'import_errors': importErrors}

            importedCount = 0
            for d_id, prod_dict in counts.items():
                db_session.query(StockRecord).filter(StockRecord.dealership_id == d_id).delete()
                for prod_name, qty in prod_dict.items():
                    rec = StockRecord(
                        dealership_id=d_id,
                        product_name=prod_name,
                        quantity=qty,
                        column_order=productOrder.get(prod_name, 0)
                    )
                    db_session.add(rec)
                    importedCount += 1

                if d_id in regionByDealer:
                    d_obj = db_session.query(Dealership).filter(Dealership.id == d_id).first()
                    if d_obj:
                        d_obj.region = regionByDealer[d_id]
        else:
            skipKeywords = [
                'sr#', 'sr.', 'sr no', 's.no', 's no', 'serial', 'dealer', 'security',
                'region', 'total', 'ttl', 'code', 'sap', 'sap code', 'odoo', 'odoo code',
                'tag', 'dealer tag', 'dealer name', 'company', 'branch', 'showroom',
                'unpaid', 'paid', 'difference', 'open', 'closed', 'cuc', 'pending', 'status'
            ]
            productCols = {}
            for col, label in enumerate(headerRow):
                labelStr = str(label).strip()
                if labelStr == '' or self.matches_any_keyword(labelStr, skipKeywords):
                    continue
                productCols[col] = labelStr

            stock_sums = {}

            for i in range(headerIndex + 1, len(rows)):
                row = rows[i]
                rowNum = i + 1
                if not any(str(c).strip() != '' for c in row):
                    continue

                dealershipName = str(row[dealerCol]).strip() if dealerCol < len(row) else ''
                if not dealershipName:
                    continue

                dealershipId = self.find_dealership_match(dealershipsByName, dealershipName)
                if not dealershipId:
                    continue

                if dealershipId in excludedStockIds:
                    continue

                if securityCol is not None and securityCol < len(row):
                    sec_str = str(row[securityCol]).strip()
                    if sec_str != '':
                        try:
                            sec_amt = float(sec_str.replace(',', ''))
                            d_obj = db_session.query(Dealership).filter(Dealership.id == dealershipId).first()
                            if d_obj:
                                d_obj.security_amount = sec_amt
                        except Exception:
                            pass

                if dealershipId not in touchedDealershipIds:
                    touchedDealershipIds.add(dealershipId)

                for col, rawProductName in productCols.items():
                    val_str = str(row[col]).strip() if col < len(row) else '0'
                    try:
                        q_float = float(val_str.replace(',', ''))
                        qty = int(q_float) if q_float < 100000 else 0
                    except Exception:
                        qty = 0
                    productName = self.normalize_stock_product_name(rawProductName)
                    key = (dealershipId, productName, col)
                    stock_sums[key] = stock_sums.get(key, 0) + qty

            if touchedDealershipIds:
                db_session.query(StockRecord).filter(StockRecord.dealership_id.in_(list(touchedDealershipIds))).delete()

            for (d_id, prod_name, col), qty in stock_sums.items():
                rec = StockRecord(
                    dealership_id=d_id,
                    product_name=prod_name,
                    quantity=qty,
                    column_order=col
                )
                db_session.add(rec)
                importedCount += 1

        db_session.commit()
        return {'success': True, 'imported_count': importedCount, 'import_errors': importErrors}

    def import_ageing_sheet(self, db_session, rows: list) -> dict:
        from api.models import AgeingRecord
        if not rows or len(rows) < 2:
            return {'success': False, 'message': 'Sheet is empty or has no data rows.'}

        dealershipsByName = self._build_dealership_map(db_session)
        headerIndex = self.find_header_row_index(rows, ['dealer'])
        headerRow = rows[headerIndex] if headerIndex < len(rows) else []

        dealerCol = self.find_column(headerRow, ['dealer'])
        dealerNameCol = self.find_column(headerRow, ['dealer name']) or dealerCol

        productCol = self.find_column(headerRow, ['product desc', 'product description']) \
            or self.find_column(headerRow, ['model name']) \
            or self.find_column(headerRow, ['product']) \
            or self.find_column(headerRow, ['model'])

        chassisCol = self.find_column(headerRow, ['chassis'])
        deliveryDateCol = self.find_column(headerRow, ['delivery date', 'deilvery date'])

        if dealerNameCol is None or chassisCol is None or deliveryDateCol is None:
            return {'success': False, 'message': 'Could not find "Dealer Name", "Chassis", and "Delivery Date" columns.'}

        touchedDealershipIds = set()
        importedCount = 0
        importErrors = []

        for i in range(headerIndex + 1, len(rows)):
            row = rows[i]
            rowNum = i + 1
            if not any(str(c).strip() != '' for c in row):
                continue

            dealershipName = str(row[dealerNameCol]).strip() if dealerNameCol < len(row) else ''
            chassis = str(row[chassisCol]).strip() if chassisCol < len(row) else ''
            deliveryDateRaw = str(row[deliveryDateCol]).strip() if deliveryDateCol < len(row) else ''

            if not dealershipName or not chassis or not deliveryDateRaw:
                continue

            dealershipId = self.find_dealership_match(dealershipsByName, dealershipName)
            if not dealershipId:
                importErrors.append(f"Row {rowNum}: Dealership \"{dealershipName}\" Not Found — Skipped.")
                continue

            deliveryDateStr = self.parse_flexible_date(deliveryDateRaw)
            if not deliveryDateStr:
                importErrors.append(f"Row {rowNum}: Could Not Parse Delivery Date \"{deliveryDateRaw}\" — Skipped.")
                continue

            try:
                delDate = datetime.strptime(deliveryDateStr, '%Y-%m-%d').date()
            except Exception:
                continue

            productName = str(row[productCol]).strip() if (productCol is not None and productCol < len(row)) else 'Vehicle'

            if dealershipId not in touchedDealershipIds:
                db_session.query(AgeingRecord).filter(AgeingRecord.dealership_id == dealershipId).delete()
                touchedDealershipIds.add(dealershipId)

            rec = AgeingRecord(
                dealership_id=dealershipId,
                product_name=productName,
                chassis_number=chassis,
                delivery_date=delDate
            )
            db_session.add(rec)
            importedCount += 1

        db_session.commit()
        return {'success': True, 'imported_count': importedCount, 'import_errors': importErrors}

    def import_stock_chassis_sheet(self, db_session, rows: list) -> dict:
        from api.models import StockChassisRecord
        if not rows or len(rows) < 2:
            return {'success': False, 'message': 'Sheet is empty or has no data rows.'}

        dealershipsByName = self._build_dealership_map(db_session)
        headerIndex = self.find_header_row_index(rows, ['dealer'])
        headerRow = rows[headerIndex] if headerIndex < len(rows) else []

        dealerCol = self.find_column(headerRow, ['dealer'])
        dealerNameCol = self.find_column(headerRow, ['dealer name']) or dealerCol
        chassisCol = self.find_column(headerRow, ['chassis'])

        if chassisCol is None:
            return {'success': False, 'message': 'Could not find "Chassis" column.'}

        touchedDealershipIds = set()
        importedCount = 0
        importErrors = []

        for i in range(headerIndex + 1, len(rows)):
            row = rows[i]
            rowNum = i + 1
            if not any(str(c).strip() != '' for c in row):
                continue

            dealershipName = str(row[dealerNameCol]).strip() if (dealerNameCol is not None and dealerNameCol < len(row)) else ''
            chassis = str(row[chassisCol]).strip() if chassisCol < len(row) else ''
            if not chassis or chassis.lower() in ('chassis', 'total'):
                continue

            dealershipId = self.find_dealership_match(dealershipsByName, dealershipName) if dealershipName else None
            if not dealershipId:
                continue

            if dealershipId not in touchedDealershipIds:
                db_session.query(StockChassisRecord).filter(StockChassisRecord.dealership_id == dealershipId).delete()
                touchedDealershipIds.add(dealershipId)

            rec = StockChassisRecord(
                dealership_id=dealershipId,
                chassis_number=chassis
            )
            db_session.add(rec)
            importedCount += 1

        db_session.commit()
        return {'success': True, 'imported_count': importedCount, 'import_errors': importErrors}

    def import_crm_sheet(self, db_session, rows: list, period_month: str) -> dict:
        from api.models import CrmParameter, CrmScore
        if not rows or len(rows) < 2:
            return {'success': False, 'message': 'Sheet is empty or has no data rows.'}

        parameters = db_session.query(CrmParameter).order_by(CrmParameter.display_order, CrmParameter.id).all()
        if not parameters:
            return {'success': False, 'message': 'No CRM Parameters defined yet — add them in CRM Parameters first.'}

        dealershipsByName = self._build_dealership_map(db_session)
        headerIndex = self.find_header_row_index(rows, ['dealer', 'company'])
        headerRow = rows[headerIndex] if headerIndex < len(rows) else []
        dealerCol = self.find_column(headerRow, ['dealer', 'dealership', 'company'])

        if dealerCol is None:
            return {'success': False, 'message': 'Could not find a "Dealer"/"Dealership"/"Company" column in header row.'}

        scoreCols = []
        for col, label in enumerate(headerRow):
            if col <= dealerCol:
                continue
            if str(label).strip() != '':
                scoreCols.append(col)

        if len(scoreCols) < len(parameters):
            return {
                'success': False,
                'message': f'Found {len(scoreCols)} score column(s) after "Dealer" but there are {len(parameters)} CRM Parameters. Check template.'
            }

        importedCount = 0
        importErrors = []

        for i in range(headerIndex + 1, len(rows)):
            row = rows[i]
            rowNum = i + 1
            if not any(str(c).strip() != '' for c in row):
                continue

            dealershipName = str(row[dealerCol]).strip() if dealerCol < len(row) else ''
            if not dealershipName:
                continue

            dealershipId = self.find_dealership_match(dealershipsByName, dealershipName)
            if not dealershipId:
                importErrors.append(f"Row {rowNum}: Dealership \"{dealershipName}\" Not Found — Skipped.")
                continue

            for pIdx, param in enumerate(parameters):
                col = scoreCols[pIdx]
                val_str = str(row[col]).strip() if col < len(row) else ''
                if val_str == '':
                    continue
                try:
                    pts = float(val_str.replace(',', ''))
                except Exception:
                    continue

                score = db_session.query(CrmScore).filter(
                    CrmScore.dealership_id == dealershipId,
                    CrmScore.crm_parameter_id == param.id,
                    CrmScore.period_month == period_month
                ).first()
                if not score:
                    score = CrmScore(
                        dealership_id=dealershipId,
                        crm_parameter_id=param.id,
                        period_month=period_month,
                        points_obtained=pts
                    )
                    db_session.add(score)
                else:
                    score.points_obtained = pts

                importedCount += 1

        db_session.commit()
        return {'success': True, 'imported_count': importedCount, 'import_errors': importErrors}

import zipfile
import xml.etree.cElementTree as ET

def fast_xlsx_parse(file_bytes):
    """Fast C-XML streaming parser for XLSX files (processes 30,000+ rows in <2 seconds)."""
    try:
        rows = []
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as zf:
            shared_strings = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                ss_tree = ET.parse(zf.open('xl/sharedStrings.xml'))
                for elem in ss_tree.getroot():
                    txt_node = elem.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                    shared_strings.append(txt_node.text if txt_node is not None else '')
                    
            sheet_file = 'xl/worksheets/sheet1.xml'
            if sheet_file not in zf.namelist():
                for name in zf.namelist():
                    if name.startswith('xl/worksheets/sheet'):
                        sheet_file = name
                        break
                        
            sheet_stream = zf.open(sheet_file)
            context = ET.iterparse(sheet_stream, events=('end',))
            for event, elem in context:
                if elem.tag.endswith('row'):
                    row_vals = []
                    for cell in elem.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                        val_type = cell.get('t')
                        val_node = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        cell_val = val_node.text if val_node is not None else ''
                        if val_type == 's' and cell_val and cell_val.isdigit():
                            idx = int(cell_val)
                            cell_val = shared_strings[idx] if idx < len(shared_strings) else cell_val
                        elif val_type == 'inlineStr':
                            is_node = cell.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                            cell_val = is_node.text if is_node is not None else ''
                        row_vals.append(str(cell_val).strip())
                    if any(c != '' for c in row_vals):
                        rows.append(row_vals)
                    elem.clear()
        if rows and len(rows) > 1:
            return rows
    except Exception:
        pass
    return None

def read_excel_rows(file_input, sheet_name: str = None) -> list:
    """Reads spreadsheet (XLSX, XLS, CSV) rows from file path string or FileStorage stream object and returns lists of strings."""
    if not file_input:
        return []

    filename = ""
    file_bytes = None

    if isinstance(file_input, str):
        if not os.path.exists(file_input):
            raise FileNotFoundError(f"File not found: {file_input}")
        filename = file_input
        with open(file_input, 'rb') as f:
            file_bytes = f.read()
    else:
        filename = getattr(file_input, 'filename', '') or ''
        file_bytes = file_input.read()
        try:
            file_input.seek(0)
        except Exception:
            pass

    if not file_bytes:
        return []

    # Try Fast C-XML Streaming Parse for XLSX first (processes 30,000+ rows in <2 seconds)
    if file_bytes.startswith(b'PK'):
        fast_rows = fast_xlsx_parse(file_bytes)
        if fast_rows:
            return fast_rows

        try:
            stream = io.BytesIO(file_bytes)
            wb = openpyxl.load_workbook(stream, data_only=True)
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                else:
                    match = [s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()]
                    sheet = wb[match[0]] if match else wb.active
            else:
                sheet = wb.active

            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(cell).strip() if cell is not None else "" for cell in row])
            if rows and any(any(c != "" for c in r) for r in rows):
                return rows
        except Exception:
            pass

    # Try XLS via xlrd if available
    try:
        import xlrd
        stream = io.BytesIO(file_bytes)
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheet = wb.sheet_by_index(0)
        rows = []
        for r_idx in range(sheet.nrows):
            row_vals = [str(sheet.cell_value(r_idx, c_idx)).strip() for c_idx in range(sheet.ncols)]
            rows.append(row_vals)
        if rows:
            return rows
    except Exception:
        pass

    # Try CSV / text decoding
    try:
        for enc in ['utf-8-sig', 'latin-1', 'cp1252']:
            try:
                text = file_bytes.decode(enc)
                reader = csv.reader(io.StringIO(text))
                rows = []
                for row in reader:
                    rows.append([str(cell).strip() if cell is not None else "" for cell in row])
                if rows and len(rows) > 1:
                    return rows
            except Exception:
                continue
    except Exception:
        pass

    return []

class DataQualityAnalyzer:
    FAKE_EMAIL_DOMAINS = ['test.com', 'example.com', 'test.test', 'asdf.com', 'abc.com', 'xyz.com', 'mail.com']

    @staticmethod
    def is_valid_email(email: str) -> bool:
        email = email.strip()
        if not email or '@' not in email:
            return False
        email_regex = r'^([a-zA-Z0-9_.+-]+)@([a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)$'
        if not re.match(email_regex, email):
            return False
        domain = email.split('@')[-1].lower()
        return domain not in DataQualityAnalyzer.FAKE_EMAIL_DOMAINS

    @staticmethod
    def is_valid_pak_phone(phone: str) -> bool:
        digits = re.sub(r'\D', '', phone)
        if not digits:
            return False
        if digits.startswith('0092'):
            digits = '0' + digits[4:]
        elif digits.startswith('92') and len(digits) == 12:
            digits = '0' + digits[2:]
        return bool(re.match(r'^03\d{9}$', digits))

    @staticmethod
    def benford_law_check(numbers: list) -> dict:
        leading_digits = []
        for n in numbers:
            try:
                val = abs(float(n))
                if val < 1:
                    continue
                leading_digits.append(int(str(int(val))[0]))
            except Exception:
                pass

        sample_size = len(leading_digits)
        if sample_size < 50:
            return None

        expected = {1: 0.301, 2: 0.176, 3: 0.125, 4: 0.097, 5: 0.079, 6: 0.067, 7: 0.058, 8: 0.051, 9: 0.046}
        observed_counts = {i: 0 for i in range(1, 10)}
        for d in leading_digits:
            if 1 <= d <= 9:
                observed_counts[d] += 1

        chi_square = 0.0
        for digit, expected_proportion in expected.items():
            expected_count = expected_proportion * sample_size
            observed = observed_counts[digit]
            chi_square += ((observed - expected_count) ** 2) / expected_count

        critical_value = 15.51
        return {
            'sample_size': sample_size,
            'chi_square': round(chi_square, 2),
            'deviates': chi_square > critical_value
        }

    @staticmethod
    def z_score_outliers(values_by_group: dict) -> dict:
        count = len(values_by_group)
        if count < 3:
            return {}

        vals = list(values_by_group.values())
        mean = sum(vals) / count
        variance = sum((v - mean) ** 2 for v in vals)
        std_dev = (variance / count) ** 0.5
        if std_dev == 0.0:
            return {}

        outliers = {}
        for key, v in values_by_group.items():
            z = (v - mean) / std_dev
            if abs(z) > 2:
                outliers[key] = round(z, 2)
        return outliers


import time
import requests
import json
from api.config import Config

class VisitReportAnalyzer:
    def __init__(self):
        self.models = ['gemini-2.5-flash', 'gemini-1.5-flash', 'gemini-flash-lite-latest', 'gemini-flash-latest']

    def analyze_weak_areas(self, context: dict) -> dict:
        prompt_lines = [
            f"You are a District Sales Manager (DSM) writing the \"Weak Areas\" section of a dealership visit report for \"{context['dealership_name']}\".",
            "Write in clear, professional English, as short paragraphs or bullet points a manager would put in an official report. Be specific and reference actual numbers throughout. If a section looks healthy, say so briefly instead of forcing a complaint.",
            "Cover the sections below in EXACTLY this order, using the specific guidance given for each:",
            "",
            "1. SALES TARGET vs ACHIEVEMENT — Compare the Grand Total Achieved against the Sales Target. If the target was NOT met, check the STOCK list below: if stock is available for the underperforming/fast-moving models, explicitly recommend that the dealership submit payment against that available stock to help achieve the monthly sales target. If the target was met or exceeded, state that briefly and move on.",
            "2. AGEING — For any vehicle aged 60+ days, urge the dealership to submit payment for those specific vehicles (name the model/chassis) to clear the ageing stock and free up tied-up capital.",
            "3. CRM & DEALERSHIP INFRASTRUCTURE — Identify weak areas among the CRM scorecard parameters below (any parameter below its max points) and recommend specific improvement actions.",
            "4. TEST DRIVE & DIGITAL CONVERSION TARGETS — Identify weak areas in the Test Drive and Digital Enquiry Conversion percentages below and recommend specific improvement actions.",
            "5. SOCIAL MEDIA & REVIEWS — Identify weak areas (followers/reviews below target, low posting frequency, poor rating, etc.) and recommend specific improvement actions.",
            "",
            "Do not mention \"PA\" or \"PB\" anywhere in your response.",
            "",
            "SALES (this month):",
        ]

        sales_target = context.get('sales_target')
        sales_grand_total = context.get('sales_grand_total')
        prompt_lines.append(f"Sales Target: {f'{sales_target:,}' if sales_target is not None else 'not set'}")
        prompt_lines.append(f"Grand Total Achieved: {f'{sales_grand_total:,}' if sales_grand_total is not None else 'not set'}")
        prompt_lines.append("")

        sales = context.get('sales', [])
        if not sales:
            prompt_lines.append("No product-wise sales data on record for this month.")
        else:
            for s in sales:
                label = SpreadsheetImportHelper.friendly_product_label(s['product_name'])
                prompt_lines.append(f"- {label}: {s['quantity']} units sold")

        prompt_lines.append("")
        prompt_lines.append("STOCK (current):")
        stock = context.get('stock', [])
        if not stock:
            prompt_lines.append("No stock data on record.")
        else:
            for s in stock:
                prompt_lines.append(f"- {s['product_name']}: {s['quantity']} units in stock")

        sec_amount = context.get('security_amount')
        if sec_amount:
            try:
                prompt_lines.append(f"Security amount held: {float(sec_amount):,}")
            except Exception:
                pass

        prompt_lines.append("")
        prompt_lines.append("AGEING (vehicles in stock 60+ days, measured against this month's last date):")
        ageing = context.get('ageing', [])
        if not ageing:
            prompt_lines.append("No vehicles aged 60+ days — nothing stuck.")
        else:
            for a in ageing:
                prompt_lines.append(f"- {a['product_name']} (Chassis {a['chassis_number']}): {a['days_aged']} days aged")

        crm = context.get('crm', [])
        general_crm = [c for c in crm if float(c.get('max_points') or 0.0) != 0.0]
        direct_result_crm = [c for c in crm if float(c.get('max_points') or 0.0) == 0.0]

        prompt_lines.append("")
        prompt_lines.append("CRM & DEALERSHIP INFRASTRUCTURE SCORECARD (this month, points obtained vs max points):")
        general_crm_has_data = any(c.get('points_obtained') is not None for c in general_crm)
        if not general_crm_has_data:
            prompt_lines.append("No CRM scorecard data on record for this month.")
        else:
            for c in general_crm:
                if c.get('points_obtained') is None:
                    continue
                below_max = " (below max)" if float(c['points_obtained']) < float(c['max_points']) else ""
                prompt_lines.append(f"- {c['parameter_name']}: {c['points_obtained']} / {c['max_points']} points{below_max}")

        prompt_lines.append("")
        prompt_lines.append("TEST DRIVE & DIGITAL CONVERSION TARGETS (this month, % achievement):")
        direct_result_has_data = any(c.get('points_obtained') is not None for c in direct_result_crm)
        if not direct_result_has_data:
            prompt_lines.append("No test drive/digital conversion data on record for this month.")
        else:
            for c in direct_result_crm:
                if c.get('points_obtained') is None:
                    continue
                below_target = " (below target)" if float(c['points_obtained']) < 100 else ""
                prompt_lines.append(f"- {c['parameter_name']}: {c['points_obtained']}% achieved{below_target}")

        prompt_lines.append("")
        prompt_lines.append("SOCIAL MEDIA & REVIEWS:")
        social = context.get('social', {})
        
        fb_tgt_str = f" (Target: {social['fb_target']:,})" if social.get('fb_target', 0) > 0 else ""
        prompt_lines.append(f"- Facebook Followers: {social.get('fb_followers', 0):,}{fb_tgt_str}")
        
        ig_tgt_str = f" (Target: {social['ig_target']:,})" if social.get('ig_target', 0) > 0 else ""
        prompt_lines.append(f"- Instagram Followers: {social.get('ig_followers', 0):,}{ig_tgt_str}")
        
        yt_tgt_str = f" (Target: {social['yt_target']:,})" if social.get('yt_target', 0) > 0 else ""
        prompt_lines.append(f"- YouTube Subscribers: {social.get('yt_subscribers', 0):,}{yt_tgt_str}")
        
        prompt_lines.append(f"- Facebook Posts/Week: {social.get('fb_posts_week', 0):,}")
        prompt_lines.append(f"- Instagram Posts/Week: {social.get('ig_posts_week', 0):,}")
        
        gr_tgt_str = f" (Target: {social['google_review_target']:,})" if social.get('google_review_target', 0) > 0 else ""
        prompt_lines.append(f"- Google Reviews: {social.get('google_review_count', 0):,} (Rating: {social.get('google_rating', 0)}/5){gr_tgt_str}")

        prompt_lines.append("")
        prompt_lines.append('Respond with ONLY this JSON, no other text: {"weak_areas": ["point 1", "point 2", "..."], "summary": "one short overall paragraph"}')
        prompt_lines.append('weak_areas should list the points IN THE ORDER of the 5 sections above (sales/stock, ageing, CRM, test drive & digital conversion, social media). If nothing significant is wrong, weak_areas should be an empty array and summary should say performance looks healthy overall.')

        body = {
            'contents': [{'parts': [{'text': "\n".join(prompt_lines)}]}],
            'generationConfig': {'response_mime_type': 'application/json'}
        }

        http_code = 0
        response_text = None
        last_message = ''
        backoff_seconds = [5, 10, 20]

        for model in self.models:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={Config.GEMINI_API_KEY}"
            
            for attempt in range(3):
                try:
                    res = requests.post(url, json=body, timeout=60)
                    http_code = res.status_code
                    response_text = res.text
                    if http_code == 200 or http_code == 429:
                        break
                except Exception as e:
                    http_code = 0
                    response_text = str(e)
                
                time.sleep(backoff_seconds[attempt] if attempt < len(backoff_seconds) else 20)

            if http_code == 200:
                break

            try:
                err_data = json.loads(response_text)
                last_message = err_data.get('error', {}).get('message') or f"HTTP {http_code}"
            except Exception:
                last_message = f"HTTP {http_code}" if http_code != 0 else f"Request Failed: {response_text}"

        if http_code != 200:
            return {'success': False, 'message': f"{last_message} (All Models Exhausted.)"}

        try:
            data = json.loads(response_text)
            text = data['candidates'][0]['content']['parts'][0]['text']
            clean_text = text.replace('```json', '').replace('```', '').strip()
            result = json.loads(clean_text)
            
            return {
                'success': True,
                'weak_areas': result.get('weak_areas', []),
                'summary': result.get('summary', '')
            }
        except Exception as e:
            return {'success': False, 'message': f"Could not parse Gemini response: {str(e)}"}

